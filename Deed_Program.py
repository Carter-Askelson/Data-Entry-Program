import sys
import openpyxl
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import PatternFill
from pathlib import Path

global row
global workbook
global sheet
global browser

titlespot = "#block-pagetitle > h1"
magnifyspot = "#block-member > div > p > strong:nth-child(2) > a"
zipspot = "#Search_zip"
addressspot = "#property-address"
searchspot = "#reports > div > div.tab-content > div.btn-group.mt-md.ml-md > div:nth-child(1) > button > strong"
owner_name_spot = "#report-block_ownership > div.panel-body > div > div:nth-child(1) > div:nth-child(1) > div.field-value"


#starts program
def main():  
    get_excel()
    get_row_max()
    set_up_site()      
    get_values()


#Finds Workbook and selects last worksheet to work in
def get_excel():
    global workbook
    global sheet
    print("In order for this programm to work the Excel File needs to have the name 'Deeds' saved to the Deed_Program Folder")
    print("A copy of the format of the Excel file is provided in the program's folder")
    print("Please note that the each run of the program will only affect THE LAST WORKSHEET in the Excel File")
    print()
    os.chdir(str(Path.cwd()))
    workbook = openpyxl.load_workbook('Deeds.xlsx')
    sheet = workbook[workbook.sheetnames[-1]]
#finds how many rows the program will run
def get_row_max():
    global sheet
    global row
    stop = 0
    row = 1
    while stop == 0:
        row = row + 1
        if sheet.cell(row=row, column=1).value == None:
            stop = 1
    #First safety measure to not charge the Safestreets HomeInfoMax account extra money
    if row > 300:
        end_program()


#gets Address and Zip Code and runs those through get_owner_name() until all rows are updated
def get_values():
    global sheet
    global row
    while row != 2:
        row -= 1
        address = sheet.cell(row=row, column=5).value
        zip_code = sheet.cell(row=row, column=6).value
    
        #Second safety measure to not charge the Safestreets HomeInfoMax account extra money
        if sheet.cell(row=row, column=3).value != None:
            end_program()
        else:
            sheet.cell(row=row, column=3).value = get_owner_name(zip_code, address)
            highlight_discrepancies()
            current_spot = browser.find_element(By.CSS_SELECTOR,magnifyspot)
            current_spot.click()
    end_program()


#opens homeinfomax.com and logs in (Make sure to update this if usernames or passwords change)
def set_up_site():
    global browser
    username = "Username Here"
    password = "Password Here"
    usernamespot = "#edit-name"
    passwordspot = "#edit-pass"
    login = "#edit-submit"
    browser = webdriver.Chrome(str(Path.cwd()) + "\\chromedriver")
    browser.get('https://www.homeinfomax.com/')
    
    current_spot = browser.find_element(By.CSS_SELECTOR,usernamespot)
    current_spot.send_keys(username)
    current_spot = browser.find_element(By.CSS_SELECTOR,passwordspot)
    current_spot.send_keys(password)
    current_spot = browser.find_element(By.CSS_SELECTOR,login)
    current_spot.click()

#using the new zip code and address from the Excel file the function searches the site and returns the name on the Deed
def get_owner_name(new_zip_code, new_address):
    current_spot = browser.find_element(By.CSS_SELECTOR,zipspot)                                    
    current_spot.send_keys(new_zip_code)
    current_spot = browser.find_element(By.CSS_SELECTOR,addressspot)                      
    current_spot.send_keys(new_address)
    current_spot = browser.find_element(By.CSS_SELECTOR,searchspot)
    current_spot.click()
    current_spot = browser.find_element(By.CSS_SELECTOR,titlespot)
  
    if current_spot.text == "Detailed Report":
        current_spot = browser.find_element(By.CSS_SELECTOR,owner_name_spot)
        return current_spot.text
    else:
        return "No hit on HIM"

#Makes the cell color red if Customer name isn't in the Deed name
def highlight_discrepancies():
    highlight_color = "ff8c8c"
    cm_name = sheet.cell(row=row, column=2).value
    cm_name = cm_name.upper()
    cm_name = cm_name.split()
    if sheet.cell(row=row, column=3).value == "No hit on HIM":
        pass
    elif len(cm_name) > 2:
        if cm_name[0] not in sheet.cell(row=row, column=3).value or cm_name[2] not in sheet.cell(row=row, column=3).value:
            sheet.cell(row=row, column=3).fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type = "solid")
    elif cm_name[0] not in sheet.cell(row=row, column=3).value or cm_name[1] not in sheet.cell(row=row, column=3).value:
        sheet.cell(row=row, column=3).fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type = "solid")

#Finishes up and ends program
def end_program():
    workbook.save("Deeds.xlsx")
    browser.quit()
    sys.exit()


main()
